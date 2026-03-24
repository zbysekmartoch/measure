"""
tento ukázkový skript načte z LAB_ROOT/Data, soubor CENY.xlsx. 
Vezme data z listu DATA z oblasti B10 až N19.
Uloží do CSV ve složce resultu (RESULT_ROOT).
Data vypíše i na stdout.
"""

import pandas as pd
from openpyxl import load_workbook
from pprint import pprint
import json
from measure.env import RESULT_ROOT, LAB_ROOT #z measure.env si naimportujeme cesty abysme mohli číst a zapisovat soubory na správná místa

print ("\npříklad 2 - načtení oblasti pomocí load_workbook a konverze do CSV\n")

# otevření souboru CENY.xlsx ze složky Data ve složce laboratoře
wb = load_workbook(LAB_ROOT + "/Data/CENY.xlsx", data_only=True)

# výběr listu DATA
ws = wb["DATA"]
data = list(ws.iter_rows(
    min_row=10,
    max_row=19,
    min_col=2,
    max_col=14,
    values_only=True
))

df = pd.DataFrame(data)

# uložíme do CSV do složky pro výstupy
df.to_csv(RESULT_ROOT+"/export z excelu.csv", index=False, encoding="utf-8-sig")

print(df)  # taky vypíšeme
